import openpyxl
import json
import tkinter as tk  # 用于生成用户界面
from tkinter import filedialog  # 用于文件对话框
import os  # 用于文件路径操作

from pathlib import Path

label = None  # 全局变量，用于存储标签引用
r = None  # 全局变量，用于存储主窗口引用

def show(main_root):

    btn = tk.Button(main_root, text="选择文件", command=select_execel_file)
    btn.pack(pady=20)
    global label  # 声明使用全局变量
    label = tk.Label(main_root, text="未选择文件", wraplength=200)
    label.pack(pady=20)
    global r  # 声明使用全局变量
    r = main_root  # 将主窗口引用存储在全局变量中

filetypes = [
        ("Excel files", "*.xlsx *.xls *.xlsm *.xlsb "),
        ("CSV files", "*.csv"),
        ("All files", "*.*")
]


def select_execel_file():
    """
    选择Excel文件
    """

    filepath = filedialog.askopenfilename(
        title="选择Excel文件", filetypes=filetypes)
    if filepath:
        label.config(text=filepath)  # 更新标签文本为选择的文件路径
        btn = tk.Button(root, text="转化成Json",
                        command=process_excel_file(filepath))
        btn.pack(pady=20)
        return filepath
    else:
        return None



def create_excel_tool_ui(parent):
    # 状态变量
    excel_path = tk.StringVar(value="未选择文件")
    output_path = tk.StringVar(value="未选择导出路径")
    result_msg = tk.StringVar(value="")
    # --- 文件选择 ---
    file_frame = tk.Frame(parent)
    file_frame.pack(pady=10)
    tk.Button(file_frame, text="选择 Excel 文件",
              command=lambda: choose_excel_file()).pack(side="left")
    tk.Label(file_frame, textvariable=excel_path, wraplength=400,
             anchor="w").pack(side="left", padx=10)
    # --- 导出路径 ---
    out_frame = tk.Frame(parent)
    out_frame.pack(pady=10)
    tk.Button(out_frame, text="选择导出路径",
              command=lambda: choose_output_folder()).pack(side="left")
    tk.Label(out_frame, textvariable=output_path, wraplength=400,
             anchor="w").pack(side="left", padx=50)

    # --- 转换按钮 ---
    convert_frame = tk.Frame(parent)
    convert_frame.pack(pady=20)
    tk.Button(convert_frame, text="转换", command=lambda: convert()
              ).pack(side="left", padx=10)
    tk.Label(convert_frame, textvariable=result_msg).pack(side="left")


    def choose_excel_file():
        path = filedialog.askopenfilename(
            title="选择 Excel 文件", filetypes=filetypes
        )
        if path:
            excel_path.set(path)
            # 获取文件名和扩展名
            file_name, file_extension = os.path.splitext(os.path.basename(path))
            folder_path = os.path.dirname(path)

            # 默认输出路径与输入文件同路径
            default_output = os.path.join(os.path.dirname(path), file_name + ".json")
            output_path.set(default_output)

    def choose_output_folder():
        floder = filedialog.askdirectory(ttitle="选择导出路径") 
        if floder and excel_path.get() != "未选择文件":
            file_name = os.path.splitext(os.path.basename(excel_path.get()))[0] + ".json"
            output_path.set(os.path.join(floder, file_name))

    def convert():
            in_path = excel_path.get()
            out_path = output_path.get()

            if in_path == "未选择文件":
                result_msg.set("❌ 请先选择 Excel 文件")
                return

            if os.path.exists(out_path):
                result_msg.set("❌ 导出文件已存在，转换失败")
                return

            try:
                process_excel_file(in_path, out_path)
                result_msg.set("✅ 转换成功")
            except Exception as e:
                result_msg.set(f"❌ 转换失败：{str(e)}")


def process_excel_file(excel_file, json_file):
    jd = []
    heads = []
    book = openpyxl.load_workbook(excel_file)
    sheet = book[u'Sheet']
    
    max_row = sheet.max_row
    max_column = sheet.max_column
    # 解析表头
    for column in range(max_column):
        heads.append(sheet.cell(1, column + 1).value)
    # 遍历每一行
    for row in range(max_row):
        if row < 2:
        	# 前两行跳过
            continue
        one_line = {}
        # 遍历一行中的每一个单元格
        for column in range(max_column): 
            k = heads[column]
            v = sheet.cell(row + 1, column + 1).value
            one_line[k] = v
        jd.append(one_line)
    book.close()
    # 将json保存为文件
    save_json_file(jd, json_file)

# 将json保存为文件
def save_json_file(jd, json_f_name):
    f = open(json_f_name, 'w', encoding='utf-8')
    txt = json.dumps(jd, indent=2, ensure_ascii=False)
    f.write(txt)
    f.close()
